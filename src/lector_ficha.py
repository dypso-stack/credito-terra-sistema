"""
Módulo: Lector de Ficha Excel Real
Crédito Terra - Banco Guayaquil
Lee la ficha FLI CT-001 completada por el analista y extrae todos los campos
para alimentar el motor de elegibilidad.
"""

import openpyxl
import warnings
from pathlib import Path
from datetime import datetime

warnings.filterwarnings('ignore', category=UserWarning)

# Mapa de celdas reales de la ficha FLI CT-001
# Verificado contra la ficha original de Banco Guayaquil
MAPA_CELDAS = {
    # Sección 0 - Cliente
    'razon_social':             ('FLI CT-001', 'C13'),
    'ruc':                      ('FLI CT-001', 'C14'),
    'actividad_principal':      ('FLI CT-001', 'C15'),
    'actividad_financiamiento': ('FLI CT-001', 'C16'),
    'sector_bg':                ('FLI CT-001', 'C17'),
    'nombre_contacto':          ('FLI CT-001', 'C18'),
    'correo_contacto':          ('FLI CT-001', 'H13'),
    'segmento':                 ('FLI CT-001', 'H14'),
    'canal_gestion':            ('FLI CT-001', 'H15'),
    'fecha_gestion':            ('FLI CT-001', 'H16'),
    'responsable':              ('FLI CT-001', 'H17'),
    'ciiu':                     ('FLI CT-001', 'H18'),

    # Sección 1 - Crédito
    'no_operacion':             ('FLI CT-001', 'C21'),
    'asesor_comercial':         ('FLI CT-001', 'C22'),
    'destino_terra':            ('FLI CT-001', 'C23'),
    'monto_credito_usd':        ('FLI CT-001', 'H21'),
    'destino_resumido':         ('FLI CT-001', 'H22'),
    'codigo_terra':             ('FLI CT-001', 'H23'),

    # Sección 2 - Proyecto
    'direccion_proyecto':       ('FLI CT-001', 'C27'),
    'coordenada_x':             ('FLI CT-001', 'C28'),
    'coordenada_y':             ('FLI CT-001', 'C29'),
    'descripcion_proyecto':     ('FLI CT-001', 'C30'),
    'etapa':                    ('FLI CT-001', 'C31'),
    'fecha_puesta_marcha':      ('FLI CT-001', 'C32'),
    'tipo_sistema':             ('FLI CT-001', 'C33'),
    'propietario_sistema':      ('FLI CT-001', 'C34'),

    # Sección 3 - Parámetros técnicos
    'energia_producida_mwh_anio':   ('FLI CT-001', 'E38'),
    'capacidad_instalada_mwp':      ('FLI CT-001', 'E39'),
    'consumo_cliente_mwh_anio':     ('FLI CT-001', 'E40'),
    'cobertura_energetica_pct':     ('FLI CT-001', 'E41'),
    'vida_util_anios':              ('FLI CT-001', 'E42'),

    # Sección 4 - Indicadores ambientales
    'factor_emision_sni':           ('FLI CT-001', 'E49'),
    'tipo_factor_emision':          ('FLI CT-001', 'E50'),
    'gei_evitadas_anual_tco2':      ('FLI CT-001', 'E51'),
    'gei_evitadas_total_tco2':      ('FLI CT-001', 'E52'),

    # Sección 5 - Financiero
    'capex_usd':                    ('FLI CT-001', 'E56'),
    'monto_credito_usd_s5':         ('FLI CT-001', 'E57'),
    'aporte_propio_usd':            ('FLI CT-001', 'E58'),
    'tarifa_electrica_usd_kwh':     ('FLI CT-001', 'E59'),
    'ahorro_anual_usd':             ('FLI CT-001', 'E60'),
    'payback_anios':                ('FLI CT-001', 'E61'),

    # Sección 6 - Elegibilidad (respuestas SI/NO en columna H)
    'uso_exclusivo_fondos':                 ('FLI CT-001', 'H65'),
    'evidencia_objetiva_proyecto':          ('FLI CT-001', 'H66'),
    'indicadores_ambientales_completos':    ('FLI CT-001', 'H67'),
    'cumple_normativa_ambiental':           ('FLI CT-001', 'H68'),
    'no_categoria_a_ifc':                   ('FLI CT-001', 'H69'),
    'no_involucra_combustibles_fosiles':    ('FLI CT-001', 'H70'),
    'no_impacto_negativo_otros_objetivos':  ('FLI CT-001', 'H71'),
    'no_lista_exclusion_saras':             ('FLI CT-001', 'H72'),
    'no_controversias_socioambientales':    ('FLI CT-001', 'H73'),
    'no_es_credito_consumo':                ('FLI CT-001', 'H74'),

    # Sección 8 - Resumen
    'observaciones':            ('FLI CT-001', 'D90'),
    'proximos_pasos':           ('FLI CT-001', 'D91'),
}

CAMPOS_ELEGIBILIDAD = [
    'uso_exclusivo_fondos', 'evidencia_objetiva_proyecto',
    'indicadores_ambientales_completos', 'cumple_normativa_ambiental',
    'no_categoria_a_ifc', 'no_involucra_combustibles_fosiles',
    'no_impacto_negativo_otros_objetivos', 'no_lista_exclusion_saras',
    'no_controversias_socioambientales', 'no_es_credito_consumo',
]


def leer_celda(ws, celda: str):
    """Lee el valor de una celda, manejando vacíos y errores de fórmula."""
    try:
        valor = ws[celda].value
        if valor is None:
            return None
        if isinstance(valor, str):
            v = valor.strip()
            return None if v in ['', '#DIV/0!', '#VALUE!', '#REF!'] else v
        return valor
    except Exception:
        return None


def convertir_booleano(valor) -> bool:
    """Convierte respuesta SI/NO de la ficha a booleano."""
    if isinstance(valor, bool):
        return valor
    if isinstance(valor, str):
        return valor.upper() in ['SI', 'SÍ', 'S', 'YES', 'TRUE', '1']
    if isinstance(valor, (int, float)):
        return bool(valor)
    return False


def convertir_numero(valor):
    """Convierte valores numéricos de la ficha, manejando coma decimal y texto."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return valor
    if isinstance(valor, str):
        v = valor.strip().replace(' ', '')
        if v == '':
            return None
        if ',' in v:  # coma decimal europea: "9,2" -> 9.2 ; "1.234,56" -> 1234.56
            v = v.replace('.', '').replace(',', '.')
        try:
            return float(v)
        except ValueError:
            return valor  # no era numérico; se conserva el original
    return valor


CAMPOS_NUMERICOS = {
    'monto_credito_usd', 'capacidad_instalada_mwp', 'energia_producida_mwh_anio',
    'consumo_cliente_mwh_anio', 'cobertura_energetica_pct', 'vida_util_anios',
    'factor_emision_sni', 'gei_evitadas_anual_tco2', 'gei_evitadas_total_tco2',
    'capex_usd', 'monto_credito_usd_s5', 'aporte_propio_usd',
    'tarifa_electrica_usd_kwh', 'ahorro_anual_usd', 'payback_anios',
}


def leer_ficha_excel(ruta_ficha: str) -> dict:
    """
    Lee una ficha FLI CT-001 completada y extrae todos los campos.

    Parámetros:
        ruta_ficha: Ruta al archivo Excel completado por el analista

    Retorna:
        dict: Todos los campos extraídos, listos para el motor de elegibilidad
    """
    ruta = Path(ruta_ficha)
    if not ruta.exists():
        raise FileNotFoundError(f"No se encontró la ficha: {ruta_ficha}")

    print(f"Leyendo ficha: {ruta.name}")
    wb = openpyxl.load_workbook(ruta, data_only=True)

    if 'FLI CT-001' not in wb.sheetnames:
        raise ValueError(f"Hoja 'FLI CT-001' no encontrada. Hojas: {wb.sheetnames}")

    ws = wb['FLI CT-001']
    caso = {}
    campos_vacios = []

    for campo, (_, celda) in MAPA_CELDAS.items():
        valor = leer_celda(ws, celda)
        if campo in CAMPOS_ELEGIBILIDAD:
            caso[campo] = convertir_booleano(valor)
        elif campo in CAMPOS_NUMERICOS:
            caso[campo] = convertir_numero(valor)
        else:
            caso[campo] = valor
        if caso[campo] is None and campo not in CAMPOS_ELEGIBILIDAD:
            campos_vacios.append(campo)

    wb.close()

    # Generar ID de caso
    no_op = caso.get('no_operacion')
    caso['id_caso'] = f"CT-{no_op}" if no_op else f"CT-{datetime.now().strftime('%Y%m%d%H%M%S')}"

    # Calcular capacidad en kWp
    cap_mwp = caso.get('capacidad_instalada_mwp')
    if isinstance(cap_mwp, (int, float)):
        caso['capacidad_kwp'] = round(cap_mwp * 1000, 2)

    # Reporte de lectura
    monto = caso.get('monto_credito_usd')
    monto_str = f"${monto:,.2f}" if isinstance(monto, (int, float)) else "No declarado"
    print(f"✅ Ficha leída correctamente")
    print(f"   Cliente:       {caso.get('razon_social') or 'No declarado'}")
    print(f"   RUC:           {caso.get('ruc') or 'No declarado'}")
    print(f"   Destino Terra: {caso.get('destino_terra') or 'No declarado'}")
    print(f"   Monto:         {monto_str}")
    print(f"   Campos vacíos: {len(campos_vacios)}/{len(MAPA_CELDAS)}")

    if campos_vacios:
        print(f"\n⚠️  Campos sin completar:")
        for c in campos_vacios:
            print(f"   - {c}")

    return caso


def procesar_ficha_completa(ruta_ficha: str) -> dict:
    """
    Flujo completo en producción:
    Lee la ficha → valida → evalúa elegibilidad → retorna dictamen.
    """
    import sys, os
    sys.path.append(os.path.dirname(__file__))
    from validador import ValidadorFicha
    from motor_elegibilidad import MotorElegibilidad

    caso = leer_ficha_excel(ruta_ficha)

    print("\nValidando datos...")
    validador = ValidadorFicha(caso)
    val = validador.validar()

    if not val['es_valido']:
        print(f"\n❌ {val['n_errores']} errores encontrados:")
        for e in val['errores']:
            print(f"   {e}")
        return {'error': True, 'validacion': val, 'caso': caso}

    if val['n_advertencias'] > 0:
        print(f"\n⚠️  {val['n_advertencias']} advertencias:")
        for a in val['advertencias']:
            print(f"   {a}")

    print("\nEvaluando elegibilidad...")
    motor = MotorElegibilidad(caso)
    dictamen = motor.emitir_dictamen()
    dictamen['detalle_criterios_dict'] = motor.resultados_criterios

    print(f"\n{'='*50}")
    print(f"DICTAMEN: {dictamen['dictamen']}")
    print(f"Criterios cumplidos: {dictamen['n_criterios_cumplidos']}/10")
    print(f"{'='*50}")

    return {'error': False, 'caso': caso, 'validacion': val, 'dictamen': dictamen}


if __name__ == "__main__":
    ruta = "data/raw/FLI__CT-001_energia_solar_fotovoltaica.xlsx"
    print("=== LECTOR DE FICHA FLI CT-001 ===")
    print("NOTA: La ficha de ejemplo está vacía.")
    print("En producción se usa la ficha completada por el analista.\n")
    resultado = procesar_ficha_completa(ruta)